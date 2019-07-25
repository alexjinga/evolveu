from openpyxl import load_workbook
from openpyxl import Workbook
my_wb = load_workbook(filename = 'Invoice-Example.xlsx')

class Object ():
    pass

class Basic_obj(dict):
    def __getattr__(self, item):
        return self[item]

def display_dictionary_line_by_line (dict):
    print (' ')
    for item in dict:
        print (item,':', dict[item] )

def display_array_line_by_line (arr):
    print (' ')
    for item in arr:
        print (item)

def make_object_dictionary_value(props_name, obj_values):
    new_obj = Basic_obj()            
    for i in range(len(props_name)):
        new_obj[props_name[i]] = obj_values[i]      
    return (new_obj)

def missing_data_in_a_row(row_data):
    for element in row_data:
        if element == None:
            return True
    return False    

def make_dictionaries_for_each_sheet(entire_sheet):
    temp_dict = {}
    duplicates = []
    missing_data = []
    counter = 0
    objects_property_names = ""
    for row in entire_sheet.values:
        # SAVING THE HEADER AS A LIST
        if counter == 0:
            objects_property_names = list(row)
            objects_property_names.pop(0)
            counter += 1
        # CREATING THE DICTIONARY
        else: 
            object_property_values = list(row)
            the_key = object_property_values.pop(0)
            the_value = make_object_dictionary_value(objects_property_names, object_property_values)
#VALIDATION 1 - step 1!!! --- checks for missing data (empty cells)
            if missing_data_in_a_row(list(row)):
                missing_data.append([the_key,the_value])
           
#VALIDATION 2 - step 1!!! --- checks for duplicate records and seves them
            if the_key in temp_dict.keys():
                duplicates.append([the_key,the_value])
            else:
                temp_dict[the_key] = the_value  
    return ([temp_dict, duplicates, missing_data])

def check_duplicate_keys (duplicate_data, current_sheet_name, exception_sheet_name, key_name):
    print(" ")
    if len(duplicate_data) != 0:
        print ("spreadsheet: <", current_sheet_name,"> has duplicated <",key_name.value,"> Keys!!!")
        if current_sheet_name == exception_sheet_name:
            print ("=== Duplicate Keys are to be expected in this file ===")
        else: 
            print ("=== Please check!!! ===")
    # displays the duplicated Keys&Data
        # display_array_line_by_line (duplicate_data) 
    # displays the duplicated keys ONLY
            key_set = []
            for item in duplicate_data:
                key_set.append(item[0])
            print (set(key_set))    
        return ([duplicate_data, current_sheet_name])
    return False

def check_missing_data (missing_data, current_sheet_name):
    print(" ")
    if len(missing_data) != 0:
        print ("Possible missing data found in spreadsheet: <", current_sheet_name,">!!!")
        print ("=== Please check!!! ===")
        display_array_line_by_line (missing_data)

def dict_value_type(dict):
    value_type_arr = []
    for item in dict:
        value_type_arr.append(type(dict[item]))
    return value_type_arr

def check_consistent_data_type(dict, sheet_name ,key_name):
    print(" ")
    diff_types = []
    default_type = [type(list(dict.keys())[0])]
    default_type.extend(dict_value_type(list(dict.values())[0])) 
    for item in dict:
        item_type = [type(item)]
        item_type.extend(dict_value_type(dict[item]))
        if item_type != default_type:
            diff_types.append([item, dict[item]])
            print("expected data type structure:", default_type)
            print("actual data type structure:  ", item_type)
            print (key_name.value,"-",item,':', dict[item])
    if len(diff_types) != 0:
        print(len(diff_types),"inconsitent data types found in spreadsheet: <", sheet_name,">!!!")
        return diff_types
    return False    

master_wb = Basic_obj() 

def read_excel_file(wb):
    ws_names = [] #will contain all the sheets names contained in the excel file
    for sheet in wb:
        ws_names.append(sheet.title)
        ws = wb.get_sheet_by_name(sheet.title) 
        result = make_dictionaries_for_each_sheet(ws)
        master_wb_value = result [0]
    #VALIDATION 1 - step 2 !!! Displaying if there is missing data    
        check_missing_data(result[2], sheet.title)
    #VALIDATION 2 - step2 !!! Displaying info if there are duplicated keys    
        check_duplicate_keys(result[1], sheet.title, 'Items Sold',ws['A1'])
    #VALIDATION 3 !!! Consistent Data Type check
        check_consistent_data_type(master_wb_value, sheet.title ,ws['A1'])
        master_wb[sheet.title] = result #dictionary containing all the sheet dictionaries
    #display the current sheet's name
        # print (sheet.title)
    #display the current sheet's associated dictionary content/data
    #display_dictionary_line_by_line (master_wb)

def find_invoice (invoice_number):
    obj = Object()
    invoices_ws = master_wb.get('Invoices')
    invoice = invoices_ws[0].get(invoice_number, None)
    if invoice == None:
        return False
    obj.client_id = invoice.get("Client ID")
    obj.invoice_date = invoice.get("Invoice Date")
    return (obj)

def find_client_info(client_id):
    clients_ws = master_wb.get('Clients')
    client = clients_ws[0].get(client_id, None)
    if client == None:
        return False
    client_name = client.get("Client Name")
    return (client_name)

def find_product(product_id):
    obj = Object()
    products_ws = master_wb.get("Product Inventory")
    product = products_ws[0].get(product_id, None)
    if product == None:
        return False
    obj.product_name = product.get("Product Name")
    obj.product_price = product.get("Price ($)")
    return obj

def find_invoice_items(invoice_number):
    arr = []
    cannot_find_item = False
    items_sold_ws = master_wb.get("Items Sold")
    item = items_sold_ws[0].get(invoice_number, None)
    if item == None:
        return False
    all_items = [item]
    for element in items_sold_ws[1]:
        if element[0] == invoice_number:
            all_items.append(element[1])
    for element in all_items:
        product_id = element.get("Product ID")
        obj = find_product(product_id)
        if obj == False:
            print("Cannot find a product with the ID: <",product_id,">")
            cannot_find_item = True
        else:
            obj.product_id = element.get("Product ID")
            obj.product_qt = element.get("Quantity")
            # print("ID: ",obj.product_id,", qty:",obj.product_qt)
            # print("prod",obj.product_name, "- $",obj.product_price )
        arr.append(obj)
    if cannot_find_item == True:
        return False
    return arr

def collect_invoice_data (invoice_number):
    obj = find_invoice(invoice_number)
    if obj == False:
        print("!!! The invoice #",invoice_number,"cannot be generated !!!")
        print("no such invoice number")
        return False
    #find client's details
    obj.client_name = find_client_info(obj.client_id)
    if obj.client_name == False:
        print("!!! The invoice #",invoice_number,"cannot be generated !!!")
        print("Cannot find client's information matching Client ID: <",obj.client_id,">")
        return False
    #find items sold and quantity
    items_quantity_n_price = find_invoice_items(invoice_number)
    if items_quantity_n_price == False:
        print("!!! The invoice #",invoice_number,"cannot be generated !!!")
        print("One or more products could not be found!")
        return False
    obj.product = items_quantity_n_price
    return obj

def create_invoice (invoice_number):
# find invoice
    obj = collect_invoice_data(invoice_number)
    if obj == False:
        print (" Invoice no.: ", invoice_number, ", cannot be generated!!!!")
    else: 
        # print ("Invoice Date:",obj.invoice_date)
        # print ("Client ID:",obj.client_id)
        # print ("Client Name:",obj.client_name)
        #for element in obj.product:
            # print (element.__dict__)
        invoice_wb = load_workbook(filename = 'Invoice-Template.xlsx')
        invoice_ws = invoice_wb.active
        invoice_ws.name = f"invoice {invoice_number}" 
    # filling the invoice header
        invoice_ws['B4'] = "Evolve U"
        invoice_ws['B5'] = "1721 29 Ave SW"
        invoice_ws['B6'] = "Calgary AB T2T-6T7"
        invoice_ws['B7'] = "403 386 5888"
        invoice_ws['B8'] = "evolve@evolveu.ca"
    # filling the rest of the invoice
        invoice_ws['F4'] = obj.invoice_date
        invoice_ws['F6'] = invoice_number
        invoice_ws['B12'] = obj.client_name
        invoice_ws['B13'] = f"Client ID: {obj.client_id}"
        invoice_ws['B14'] = "Calgary AB"
        invoice_ws['B15'] = "403 123 4567"
        invoice_ws['B16'] = "e-mail@email.com"
        for i in range(len(obj.product)):
            if ((type(obj.product[i].product_price) != int) and (type(obj.product[i].product_price) != float)) or (type(obj.product[i].product_qt) != int):
                print (" ")
                print("Cannot generate the invoice !!!! the price or the quantity is not a number")   
                return False
            cell = 20 + i
            invoice_ws[f'B{cell}'] = obj.product[i].product_name
            invoice_ws[f'D{cell}'] = obj.product[i].product_qt
            invoice_ws[f'E{cell}'] = obj.product[i].product_price
    #save the new excel file/invoice
        invoice_wb.save (filename = f"invoice no.{invoice_number}-{obj.client_name}.xlsx")

def cross_reference_validation ():
    faulty_invoices = []
    invoices_ws = master_wb.get('Invoices')
    invoices = invoices_ws[0]
    # print (invoices_ws[0])
    for item in invoices:
        # print (item,':', invoices[item] )
        if collect_invoice_data(item) == False: 
            faulty_invoices.append(item)
    print(" ")
    print("=== CROSS REFERENCE VALIDATION SUMMARY ===")
    if len(faulty_invoices) == 0:
        print ("0 errors have been found!!!!")
    else: 
        print (len(faulty_invoices)," invoices could not be generated")
        print ("See bellow the list:")
        display_array_line_by_line(faulty_invoices)

def calculate_total (obj):
    total = 0
    for item in obj.product:
        if (type(item.product_price) != int) and (type(item.product_price) != float):
            print(f"{item.product_name} price is not a number <{item.product_price}>! Cannot calculate the total!")   
            return False
        elif (type(item.product_qt) != int):
            print(f"{item.product_name} quantity is not a number <{item.product_qt}>! Cannot calculate the total!") 
            return False
        total += (item.product_price * item.product_qt)
    return total

def total_by_client (client_id):
    can_calculate_total = True
    client_name = find_client_info(client_id)
    if client_name == False:
        print("Cannot find client's information matching Client ID: <",client_id,">")
        return False
    clients_invoices = []
    total_spent = 0
    invoices_ws = master_wb.get('Invoices')
    invoices = invoices_ws[0]
#check how many invoices does the client have
    for item in invoices:
        if invoices[item].get("Client ID") == client_id:
            clients_invoices.append(item)
    # print(clients_invoices)
    if len(clients_invoices) == 0:
        print ("Client: ", client_name, "has not bought anything yet")
        return True
#calculate the total spent
    for item in clients_invoices:
        obj = collect_invoice_data(item)
        amount_spent = calculate_total(obj)
        if amount_spent == False:
            can_calculate_total = False
        else:
            total_spent += amount_spent
    print(" ")
    if can_calculate_total == True:
        print ("Client: ", client_name, "has spent a total of: $",round(total_spent))
    else:
        print ("Cannot calculate <", client_name, "> total spending")
    

read_excel_file(my_wb)

cross_reference_validation()

total_by_client ('0001')

create_invoice('00059')